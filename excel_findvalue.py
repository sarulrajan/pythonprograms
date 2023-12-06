import pandas as pd
import numpy as np
import requests
from openpyxl import load_workbook


wb1 = load_workbook(r'pip jobcard kodiyakadu.xlsx')  
wb2 = load_workbook(r'pip format kodiyakadu.xlsx')

jobcard = wb1.active  
pip = wb2.active

#b2-b6 w2
#k1-k5 w1
x = 0;
k = 0;


for c1 in range(3,70):
   x = pip.cell(c1, 8)
      
   print("xvalue")
   print(x.value)
   
   print("hi")
   if (x.value != None):
      print(x.value)
      for f2 in range(3,926):
         k = jobcard.cell(f2, 5)
         print("kvalue")
         print(k.value)
         if (x.value == k.value):
             pip.cell(c1, 9).value = jobcard.cell(f2, 3).value
             print("equal")
             print(jobcard.cell(f2, 3).value)
             #print(x.value)
             #print(k.value)
             print(f2)
             wb2.save('C:/Users/User/AppData/Local/Programs/Python/Python38/pip format kodiyakadu.xlsx')
             x = 0;
             k = 0;
             print(" one record completed")
             break;
         else:
             print('not equal')
   else:
      print("No value")
   