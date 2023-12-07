# to delete allocation in nrega


from selenium.webdriver.support.select import Select
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from tkinter import *
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver import ActionChains






options = webdriver.ChromeOptions()
options.add_experimental_option('debuggerAddress', 'localhost:8989')
browser = webdriver.Chrome(r'C:\dri\chromedriver.exe',options=options)



browser.get('https://nregade1.nic.in/netnrega/delWrkAlloc.aspx');

#list of numbers
path = "C:/Users/User/AppData/Local/Programs/Python/Python38/kodi.xlsx";

wb = load_workbook(path);

ws = wb.active

sno = 1;
row1 = 1;
pts = ws.cell(row1,11).value;


select = Select(browser.find_element_by_id('ctl00_ContentPlaceHolder1_ddlpanchayat_code'));
select.select_by_visible_text(pts);
print(" panchayat selected");
    
while sno != None:
    id = ws.cell(row1,4).value;
    sno = ws.cell(row1, 2).value;

    browser.find_element_by_id('ctl00_ContentPlaceHolder1_txtRegSearch').clear();
    browser.find_element_by_id('ctl00_ContentPlaceHolder1_txtRegSearch').send_keys(id);
    time.sleep(2);

    act = ActionChains(browser)
    act.send_keys(Keys.TAB).perform()
    select = Select(browser.find_element_by_id('ctl00_ContentPlaceHolder1_ddlRegistration'));
    time.sleep(2);
    select.select_by_index(1);

    browser.find_element_by_id('ctl00_ContentPlaceHolder1_GridView1_ctl02_chkAllocate')
    browser.find_element_by_id('ctl00_ContentPlaceHolder1_GridView1_ctl02_chkAllocate').click();
    time.sleep(2);
        
    browser.find_element_by_id('ctl00_ContentPlaceHolder1_cmdUpdate')
    browser.find_element_by_id('ctl00_ContentPlaceHolder1_cmdUpdate').click();
        
    ws.cell(row1, 12).value = "deleted";
    wb.save('C:/Users/User/AppData/Local/Programs/Python/Python38/kodi.xlsx');
    
    row1 += 1

